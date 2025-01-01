"""
====================
pkct_report.py
Report Generation For PKCT Tool
Author: Shubham
===================
 
"""
import os
import re
import subprocess
import difflib
import requests
import csv
from git import Repo, exc
from git.exc import GitCommandError
from datetime import datetime
from bs4 import BeautifulSoup
import shutil
import settings
import django
import argparse
from uuid import UUID
from urllib.parse import quote
from datetime import datetime
import sys
import logging
import json
import base64
import settings
from urllib.parse import quote
from django.utils.timezone import make_naive
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from pkct_main import *
from hmi.CVEHMI.pkct_report_functions import *

# Initialize Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'settings') 
django.setup()

from serverapp.models import CVE, CVEDescription,MatchString, CPEMatchInNode, CVEReference, MatchString, CPEMatch, CPEMatchInNode, CVSSMetricV2, CVSSMetricV31, CVEWeakness

LOGS_DIR = os.path.join(settings.PROJECT_DIR,'hmi','CVEHMI','hmiapp','media','logs')
HMI_REPORT = os.path.join(settings.PROJECT_DIR,'hmi','CVEHMI','hmiapp','media','reports','download')
os.makedirs(LOGS_DIR, exist_ok=True)
os.makedirs(HMI_REPORT, exist_ok=True)


# Configure logging
def setup_logging(scan_id):
    log_file_path = os.path.join(LOGS_DIR, f"{scan_id}_scan.log")
    logging.basicConfig(
        filename=log_file_path,
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )

logger = logging.getLogger()

DICTIONARY_PATH = settings.DICTIONARY_PATH
PKCT_REPORT_PATH = settings.PKCT_REPORT_PATH
DIFF_SIMILARITY_THRESHOLD = 1.0  # Threshold for diffs similarity comparison
# Path to the config.json file (adjust based on your project structure)
CONFIG_FILE_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', 'config.json'))

#########################################################################################################################
# Fetch CVE Data from database

def fetch_cve_data(cve_ids):
    cve_data = defaultdict(lambda: defaultdict(set))
    for cve_id in cve_ids:
        try:
            cve = CVE.objects.get(id=cve_id)
            descriptions = CVEDescription.objects.filter(cve=cve, lang='en')  # Filter for English descriptions only
            cvss_v2_metrics = CVSSMetricV2.objects.filter(cve=cve)
            cvss_v31_metrics = CVSSMetricV31.objects.filter(cve=cve)
            weaknesses = CVEWeakness.objects.filter(cve=cve)
            references = CVEReference.objects.filter(cve=cve)

            for desc in descriptions:
                # Convert timezone-aware datetimes to naive datetimes
                published_naive = make_naive(cve.published)
                last_modified_naive = make_naive(cve.last_modified)

                # Collect data
                cve_data[cve.id]["source_identifier"].add(cve.source_identifier)
                cve_data[cve.id]["published"].add(published_naive.strftime("%Y-%m-%d %H:%M:%S"))
                cve_data[cve.id]["last_modified"].add(last_modified_naive.strftime("%Y-%m-%d %H:%M:%S"))
                cve_data[cve.id]["vuln_status"].add(cve.vuln_status)
                cve_data[cve.id]["description_lang"].add(desc.lang)
                cve_data[cve.id]["description_value"].add(desc.value)

            for cvss_v2 in cvss_v2_metrics:
                cve_data[cve.id]["cvss_v2_source"].add(cvss_v2.source)
                cve_data[cve.id]["cvss_v2_version"].add(cvss_v2.version)
                cve_data[cve.id]["cvss_v2_vector_string"].add(cvss_v2.vector_string)
                cve_data[cve.id]["cvss_v2_access_vector"].add(cvss_v2.access_vector)
                cve_data[cve.id]["cvss_v2_base_score"].add(str(cvss_v2.base_score))
                cve_data[cve.id]["cvss_v2_exploitability_score"].add(str(cvss_v2.exploitability_score))
                cve_data[cve.id]["cvss_v2_impact_score"].add(str(cvss_v2.impact_score))

            for cvss_v31 in cvss_v31_metrics:
                cve_data[cve.id]["cvss_v31_source"].add(cvss_v31.source)
                cve_data[cve.id]["cvss_v31_version"].add(cvss_v31.version)
                cve_data[cve.id]["cvss_v31_vector_string"].add(cvss_v31.vector_string)
                cve_data[cve.id]["cvss_v31_attack_vector"].add(cvss_v31.attack_vector)
                cve_data[cve.id]["cvss_v31_privileges_required"].add(cvss_v31.privileges_required)
                cve_data[cve.id]["cvss_v31_base_score"].add(str(cvss_v31.base_score))
                cve_data[cve.id]["cvss_v31_exploitability_score"].add(str(cvss_v31.exploitability_score))
                cve_data[cve.id]["cvss_v31_impact_score"].add(str(cvss_v31.impact_score))

            for weakness in weaknesses:
                weakness_info = f"Source: {weakness.source}, Type: {weakness.type}, Description: {weakness.description}"
                cve_data[cve.id]["weaknesses"].add(weakness_info)

            for reference in references:
                reference_info = f"URL: {reference.url}, Source: {reference.source}, Tags: {reference.tags}"
                cve_data[cve.id]["references"].add(reference_info)

        except CVE.DoesNotExist:
            print(f"CVE ID {cve_id} not found in the database.")
            logger.info(f"CVE ID {cve_id} not found in the database.")
    return cve_data

#############################################################################################################
# Save To Excel Report

def save_to_excel(cve_data, results, output_file, sections=None, result_file=None):
    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Define the new column headers in the desired order
    headers = [
        "CVE Id", "Patch Status", "Status Detail", "Patch File URL",
        "Source Identifier", "Published", "Last Modified", "Vulnerability Status"
    ]

    # Determine available sections if none are provided
    if sections is None:
        sections = [
            "Description", "CVSSV2", "CVSSV3.1", "Weaknesses", "References"
        ]

    # Add additional headers based on the sections argument
    if "Description" in sections:
        headers.append("Description")
    if "CVSSV2" in sections:
        headers.extend([
            "CVSS v2 Source", "CVSS v2 Version", "CVSS v2 Vector String",
            "CVSS v2 Access Vector", "CVSS v2 Base Score", "CVSS v2 Exploitability Score", "CVSS v2 Impact Score"
        ])
    if "CVSSV3.1" in sections:
        headers.extend([
            "CVSS v3.1 Source", "CVSS v3.1 Version", "CVSS v3.1 Vector String",
            "CVSS v3.1 Attack Vector", "CVSS v3.1 Privileges Required",
            "CVSS v3.1 Base Score", "CVSS v3.1 Exploitability Score", "CVSS v3.1 Impact Score"
        ])
    if "Weaknesses" in sections:
        headers.append("Weaknesses")
    if "References" in sections:
        headers.append("References")

    ws.append(headers)

    # Create a dictionary for patch statuses from results
    patch_status_dict = {cve_id: status for cve_id, status in results}

    # Read the first commit IDs from the result file
    commit_map = {}
    if result_file:  # Check if result_file is provided
        with open(result_file, 'r') as file:
            for line in file:
                commit_id, patch_filename = line.strip().split(' -> ')
                cve_id = patch_filename.split('_')[0]  # Extract CVE ID from the patch filename
                if cve_id not in commit_map:  # Only keep the first commit ID
                    commit_map[cve_id] = commit_id

    # Initialize a dictionary to keep track of maximum width for each column
    column_widths = {i + 1: len(str(header)) for i, header in enumerate(headers)}

    # Append the data to the worksheet
    for cve_id in sorted(cve_data.keys()):
        row = [
            cve_id,
            patch_status_dict.get(cve_id, "Unknown"),  # Get Patch Status from results
            "",  # Placeholder for Status Detail
            "",  # Placeholder for Patch File URL
            ", ".join(cve_data[cve_id]["source_identifier"]),
            ", ".join(cve_data[cve_id]["published"]).split(" ")[0],  # Trim time part
            ", ".join(cve_data[cve_id]["last_modified"]).split(" ")[0],  # Trim time part
            ", ".join(cve_data[cve_id]["vuln_status"]),
        ]

        # Extract Status Detail if available
        status_detail = ""
        if "(" in patch_status_dict.get(cve_id, ""):
            status, detail = patch_status_dict[cve_id].split("(", 1)
            status_detail = detail.strip(" )")  # Remove parentheses and whitespace
            row[1] = status.strip()  # Update Patch Status without detail
        row[2] = status_detail  # Update Status Detail

        # Add the Patch File URL for the first commit ID
        commit_id = commit_map.get(cve_id)
        if commit_id:
            row[3] = f"https://git.kernel.org/pub/scm/linux/kernel/git/stable/linux.git/patch/?id={commit_id}"
        else:
            row[3] = "Not Available"  # If no commit ID found

        # Add additional data based on the sections argument
        if "Description" in sections:
            row.append(", ".join(cve_data[cve_id].get("description_value", [])))
        if "CVSSV2" in sections:
            row.extend([
                ", ".join(cve_data[cve_id].get("cvss_v2_source", [])),
                ", ".join(cve_data[cve_id].get("cvss_v2_version", [])),
                ", ".join(cve_data[cve_id].get("cvss_v2_vector_string", [])),
                ", ".join(cve_data[cve_id].get("cvss_v2_access_vector", [])),
                ", ".join(cve_data[cve_id].get("cvss_v2_base_score", [])),
                ", ".join(cve_data[cve_id].get("cvss_v2_exploitability_score", [])),
                ", ".join(cve_data[cve_id].get("cvss_v2_impact_score", [])),
            ])
        if "CVSSV3.1" in sections:
            row.extend([
                ", ".join(cve_data[cve_id].get("cvss_v31_source", [])),
                ", ".join(cve_data[cve_id].get("cvss_v31_version", [])),
                ", ".join(cve_data[cve_id].get("cvss_v31_vector_string", [])),
                ", ".join(cve_data[cve_id].get("cvss_v31_attack_vector", [])),
                ", ".join(cve_data[cve_id].get("cvss_v31_privileges_required", [])),
                ", ".join(cve_data[cve_id].get("cvss_v31_base_score", [])),
                ", ".join(cve_data[cve_id].get("cvss_v31_exploitability_score", [])),
                ", ".join(cve_data[cve_id].get("cvss_v31_impact_score", [])),
            ])
        if "Weaknesses" in sections:
            row.append(", ".join(cve_data[cve_id].get("weaknesses", [])))
        if "References" in sections:
            row.append(", ".join(cve_data[cve_id].get("references", [])))

        ws.append(row)

        # Update column widths based on the content length
        for i, cell_value in enumerate(row, 1):
            if cell_value:  # Only check non-empty cells
                # Calculate the width with some padding and limit maximum width
                width = min(len(str(cell_value)) + 2, 100)  # +2 for padding, max 100
                column_widths[i] = max(column_widths[i], width)

    # Apply the calculated column widths
    for column, width in column_widths.items():
        column_letter = get_column_letter(column)
        ws.column_dimensions[column_letter].width = width

    # Save the workbook
    wb.save(output_file)
    print(f"Data saved to Excel File: {output_file}")
    logger.info(f"Data saved to Excel File: {output_file}")


###########################################################################################

def load_blacklist(blacklist_file):
    """
    Load CVE IDs from the provided blacklist file.
    
    :param blacklist_file: Path to the CSV or XLSX file.
    :return: Set of CVE IDs to exclude.
    """
    cve_ids = set()
    try:
        if blacklist_file.endswith('.csv'):
            df = pd.read_csv(blacklist_file, header=None)
        elif blacklist_file.endswith('.xlsx'):
            df = pd.read_excel(blacklist_file, header=None)
        else:
            logger.error("Unsupported file format. Please provide a CSV or XLSX file.")
            return cve_ids
        
        # Assuming CVE IDs are in the first column
        cve_ids = set(df[0].dropna().astype(str).tolist())
    except Exception as e:
        logger.error(f"Error reading blacklist file: {e}")
    
    return cve_ids

#############################################################################################

# Filter CVE IDs on Patch Status

def filter_patch_status_in_excel(excel_file_path, patch_status_terms):
    # Read the Excel file
    df = pd.read_excel(excel_file_path)

    # Convert patch status terms to lowercase for case-insensitive comparison
    patch_status_terms_lower = [term.lower() for term in patch_status_terms]

    # Filter rows where the Patch Status column contains any of the specified terms
    df_filtered = df[df['Patch Status'].str.lower().isin(patch_status_terms_lower)]

    # Write the filtered DataFrame back to the same Excel file
    df_filtered.to_excel(excel_file_path, index=False)

    print(f"Filtered results have been written back to {excel_file_path}")

#############################################################################################

def main():
    check_ssh_agent()
    parser = argparse.ArgumentParser(description="Kernel CVE Check Tool")

    # Repository and branch arguments
    parser.add_argument('-gk', '--githubkernel', help='GitHub link to user kernel repository', required=False)
    parser.add_argument('-gb', '--githubbranch', help='Branch for the GitHub user kernel repository', required=False)
    parser.add_argument('-db', '--dotbranch', help='Upstream dot kernel branch', required=False)
    parser.add_argument('-u', '--user',help='original user name', required=False)
    parser.add_argument('-pname', '--projectname', help='project name', required=False)
    parser.add_argument('-ub', '--upstreambranch', help='upstream kernel branch', required=False)
    parser.add_argument('-build', '--buildpath', help='Build file ', required=True)
    parser.add_argument("-p", "--package_name", help="The name of the package to search.", required=False)
    parser.add_argument("-v", "--version", help="The version of the package to search.", required=False)
    parser.add_argument("-m", "--manifest_filename", help="The filename of the manifest located in the 'sample_manifest' folder.")
    parser.add_argument("-s", "--sections", nargs='*', help="Additional sections to include in the report (e.g., Description, CVSSV2, CVSSV3.1, Weaknesses, References)")
    parser.add_argument('--filter', nargs='+', help='Filters to apply (e.g. PublishedDate 20-05-2022 CVSSV2Base 2.0 CVSSV2Exploitability 1.5 CVSSV2Impact 1.5 CVSSV3.1Base 1.6 CVSSV3.1Exploitability 2.2 CVSSV3.1Impact 2.2)')
    parser.add_argument('-bl', '--blacklist', help='Path to the CSV or XLSX file containing CVE IDs to exclude', required=False)
    parser.add_argument('--username', help='request.user.username', required=False, default= "default_username")
    parser.add_argument('--project_id', help='project_id', required=False, default= "project_id")
    parser.add_argument('--scan_id', help='scan id', required=False, default= "scan_id")
    parser.add_argument('--report_name', help='The name of the report to generate')

    os.system('eval $(ssh-agent)')
    os.system('ssh-add /root/.ssh/id_rsa')
    print('\n')
    print('\n')
    args = parser.parse_args()

    username = args.username
    scan_id = args.scan_id
    project_id = args.project_id

    setup_logging(scan_id)
    print("Execution command:", ' '.join(os.sys.argv))    

    # Ensure required parameters are provided
    if not args.buildpath:
        print("Error: The build path is required.")
        logger.error("Error: The build path is required.")
        exit(1)

    # Load paths from config.json
    config_data = load_paths_from_config()

    if config_data:
        # Access specific paths stored in the config file
        download_path = config_data.get('download_dir', None)
        upload_path = config_data.get('upload_dir', None)
        working_dir = config_data.get('working_dir', None)

        # Check if the paths exist and display the result
        if download_path:
            check_path_exists(download_path, "Absolute Path 1")
        else:
            print("Download Path is not defined in config.json.")
            logger.error("Download Path is not defined in config.json.")
            exit(1)

        if upload_path:
            check_path_exists(upload_path, "Absolute Path 2")
        else:
            print("Upload Path is not defined in config.json.")
            logger.error("Upload Path is not defined in config.json.")
            exit(1)
        if working_dir:
            check_path_exists(working_dir, "Absolute Path 3")
        else:
            print("Working directory is not defined in config.json.")
            logger.error("Working directory is not defined in config.json.")
            exit(1)

    else:
        print("No paths available in config.json. Please first write the path for temporary files")
        logger.error("No paths available in config.json. Please first write the path for temporary files")
        exit(1)

    user_project_dir = None
    # Create the directory with the user name and project name under the download path
    if args.user and args.projectname and download_path:
        user_project_dir = os.path.join(download_path, args.user, args.projectname)
    

        # Check if the directory exists; if not, create it
        if not os.path.exists(user_project_dir):
            os.makedirs(user_project_dir)
            print(f"Directory '{user_project_dir}' created successfully.")
            logger.info(f"Directory '{user_project_dir}' created successfully.")
        else:
            print(f"Directory '{user_project_dir}' already exists.")
            logger.info(f"Directory '{user_project_dir}' already exists.")
    else:
        print("Download path, user, or project name is not defined.")
        logger.error("No paths available in config.json. Please first write the path for temporary files")
        exit(1)

    download_path = config_data.get('download_dir', None)
    upload_path = config_data.get('upload_dir', None)
    working_dir = config_data.get('working_dir',None)

    build_file_path = args.buildpath

    # Use the provided arguments or fall back to default configuration
    USER_KERNEL_REPO_URL = args.githubkernel
    USER_KERNEL_BRANCH_NAME = args.githubbranch
    STABLE_USER_KERNEL_REPO_URL = 'https://git.kernel.org/pub/scm/linux/kernel/git/stable/linux-stable.git'
    UPSTREAM_REPO_URL = "https://git.kernel.org/pub/scm/linux/kernel/git/torvalds/linux.git"
    UPSTREAM_BRANCH_NAME = args.upstreambranch or "master"
    STABLE_USER_KERNEL_BRANCH_NAME = args.dotbranch

    user_kernel_repo_path = None
    # Step 1: Clone and checkout
    if USER_KERNEL_REPO_URL and USER_KERNEL_BRANCH_NAME:
        repo_dir_name = os.path.splitext(os.path.basename(USER_KERNEL_REPO_URL))[0]
        target_folder = 'userkernel'
        user_kernel_repo_path, user_log_path = branch_clone_and_checkout(USER_KERNEL_REPO_URL, USER_KERNEL_BRANCH_NAME, repo_dir_name, target_folder, upload_path, user_project_dir)
        if not user_kernel_repo_path:
            print("Failed to prepare local repository. Exiting. Project link or project branch name is invalid")
            logger.error("Failed to prepare local repository. Exiting. Project link or project branch name is invalid ")
            exit(1)
            

    stable_repo_path = None

    if STABLE_USER_KERNEL_BRANCH_NAME and STABLE_USER_KERNEL_REPO_URL:
        repo_dir_name = 'dot-stable-kernel'
        target_folder = 'linux-stable'
        logs_dir_name = "stable-logs"
        stable_repo_path, stable_log_path = clone_and_checkout(STABLE_USER_KERNEL_REPO_URL, STABLE_USER_KERNEL_BRANCH_NAME, repo_dir_name, target_folder, logs_dir_name, upload_path, download_path)
        if not stable_repo_path:
            print("Failed to clone stable repo")
            logger.error("Failed to clone stable repo")
            exit(1)

    upstream_repo_path = None

    if UPSTREAM_REPO_URL and UPSTREAM_BRANCH_NAME:
        repo_dir_name = "upstream_kernel"
        target_folder = 'mainline'
        logs_dir_name = "upstream-logs"
        upstream_repo_path, upstream_log_path = clone_and_checkout(UPSTREAM_REPO_URL, UPSTREAM_BRANCH_NAME, repo_dir_name, target_folder, logs_dir_name, upload_path, download_path)
        if not upstream_repo_path:
            print("Failed to clone stable repo")
            logger.error("Failed to clone linux mainline repo")
            exit(1)

    pack = args.package_name or 'kernel'

    if pack and args.version:
        # Strip any leading or trailing spaces from the package name and version
        package_name = args.package_name or 'kernel'
        version = args.version

        # Search directly using the provided package name and version
        manifest_entries = [(package_name, version)]

    elif args.manifest_filename:
        # Load manifest from the provided filename
        manifest_file_path = args.manifest_filename
        manifest_entries = load_manifest(manifest_file_path)

    else:
        # No arguments provided, display warning and exit
        print("Warning: Please provide either a manifest file name or both a package name and version.")
        logger.warning("Warning: Please provide either a manifest file name or both a package name and version.")
        exit(1)

    # Path to the CSV file containing the package dictionary
    csv_dict_file_path = os.path.join(DICTIONARY_PATH)

    cve_ids = []

    if args.manifest_filename:
        cve_ids= accumulate_cve_ids_manifest(manifest_entries,csv_dict_file_path)
        if len(cve_ids)==0:
            logger.error("No kernel package name found in manifest")
    if args.version:
        cve_ids = accumulate_cve_ids(manifest_entries, csv_dict_file_path)

    results = []
    output_dir = "patch_files"
    result_file = "commit_map_patch_result.txt"

    output_dir = os.path.join(download_path, "patch_files")
    os.makedirs(output_dir, exist_ok=True)

    result_file = os.path.join(working_dir, "commit_map_patch_result.txt")

    blacklist_ids = set()
    if args.blacklist:
        blacklist_ids = load_blacklist(args.blacklist)

    # Apply filters if provided

    patch_status_filter = []    
    published_date_filter = None
    cvss_v2_base_filter = None
    cvss_v2_exploitability_filter = None
    cvss_v2_impact_filter = None
    cvss_v31_base_filter = None
    cvss_v31_exploitability_filter = None
    cvss_v31_impact_filter = None
    

    if args.filter:
        for i in range(len(args.filter)):     
            if args.filter[i] == "PatchStatus":
                patch_status_filter = []
                for status in args.filter[i + 1:]:
                    if status not in ["PublishedDate", "CVSSV2Base", "CVSSV2Exploitability", "CVSSV2Impact", "CVSSV3.1Base", "CVSSV3.1Exploitability", "CVSSV3.1Impact"]:
                        patch_status_filter.append(status)
                    else:
                        break    
            elif args.filter[i] == "PublishedDate":
                published_date_filter = args.filter[i + 1]  # Get the date
            elif args.filter[i] == "CVSSV2Base":
                cvss_v2_base_filter = float(args.filter[i + 1])  # Get the CVSS V2 Base Score filter value
            elif args.filter[i] == "CVSSV2Exploitability":
                cvss_v2_exploitability_filter = float(args.filter[i + 1])  # Get the CVSS V2 Exploitability Score filter value
            elif args.filter[i] == "CVSSV2Impact":
                cvss_v2_impact_filter = float(args.filter[i + 1])  # Get the CVSS V2 Impact Score filter value
            elif args.filter[i] == "CVSSV3.1Base":
                cvss_v31_base_filter = float(args.filter[i + 1])  # Get the CVSS V3.1 Base Score filter value
            elif args.filter[i] == "CVSSV3.1Exploitability":
                cvss_v31_exploitability_filter = float(args.filter[i + 1])  # Get the CVSS V3.1 Exploitability Score filter value
            elif args.filter[i] == "CVSSV3.1Impact":
                cvss_v31_impact_filter = float(args.filter[i + 1])  # Get the CVSS V3.1 Impact Score filter value

    for cve_id in cve_ids:
        references = get_filtered_references(cve_id)
        if references:
            for reference in references:
                url, source = reference  # Unpack the tuple to get the URL and its source
                print(f"\nProcessing URL: {url} (Source: {source})")
                commit_id = extract_commit_id_from_url(url)
                # Generate patch file and Save the results to a text file
                generate_patch_files(upstream_repo_path, commit_id, cve_id, output_dir, result_file)

    base_dir = os.getcwd()  # Get the current working directory

    # Define the directory for patch files and the result file path
    patch_files_path = os.path.join(download_path, 'patch_files')
    result_file_path = os.path.join(working_dir, 'commit_map_patch_result.txt')

    # Step 2: Get filtered references
    for cve_id in cve_ids:
        references = get_filtered_references(cve_id)
        if not references:
            status = "CHECK-MANUALLY   (No Patch url available)"
            print(f"{cve_id}    No Patch url available")

        else:
            # Step 3: Process each reference
            status = process_references(user_log_path, stable_log_path, upstream_log_path, user_kernel_repo_path, upstream_repo_path, references, cve_id, build_file_path, patch_files_path, result_file_path)
            print(f"{cve_id}    {status}")
        results.append((cve_id, status))

    # Filter out blacklisted CVE IDs from results
    results = [(cve_id, status) for cve_id, status in results if cve_id not in blacklist_ids]        

    # Generate a timestamped filename for the results
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file_path = os.path.join(PKCT_REPORT_PATH, f'pkct_results_{timestamp}.txt')    

    # Generate the Report
    generate_report(output_file_path, results)
    logger.info((f"Results have been written to {output_file_path}"))
    print(f"Results have been written to {output_file_path}")
    
    # Write the patch file URLs for CVE IDs into the results file
    write_patch_urls_to_results(output_file_path, result_file_path)

    # Fetch CVE data and save to Excel
    cve_data = fetch_cve_data(cve_ids)

    # Create a patch status dictionary from results
    patch_status_dict = {cve_id: status for cve_id, status in results}

    # Filter out blacklisted CVE IDs from cve_data
    cve_data = {cve_id: data for cve_id, data in cve_data.items() if cve_id not in blacklist_ids}

    # Apply filters to cve_data before saving to Excel
    if patch_status_filter or published_date_filter or cvss_v2_base_filter is not None or cvss_v2_exploitability_filter is not None or cvss_v2_impact_filter is not None or cvss_v31_base_filter is not None or cvss_v31_exploitability_filter is not None or cvss_v31_impact_filter is not None :
        filtered_cve_data = {}

        for cve_id, data in cve_data.items():
            # Ensure published date is a string (take the first one if it's a set)
            published_dates = data.get('published', set())
            if isinstance(published_dates, set):
                published_date = next(iter(published_dates), None)  # Get the first element or None
            else:
                published_date = published_dates  # Already a string

            # Check Patch Status filter
            if patch_status_filter and cve_id in patch_status_dict:
                if not any(status in patch_status_dict[cve_id] for status in patch_status_filter):
                    continue  # Skip if patch status does not match
            
            # Check Published Date filter
            if published_date_filter and published_date:
                try:
                    published_date_dt = datetime.strptime(published_date, "%Y-%m-%d %H:%M:%S")
                    filter_date = datetime.strptime(published_date_filter, "%d-%m-%Y")
                    if published_date_dt < filter_date:
                        continue  # Skip if published date is less than the filter date
                except ValueError:
                    print(f"Invalid date format for CVE ID {cve_id}: {published_date}")
                    logger.error(f"Invalid date format for CVE ID {cve_id}: {published_date}")
                    continue  # Skip if the date format is invalid

            # Check CVSS V2 Base Score filter
            if cvss_v2_base_filter is not None:
                cvss_v2_base_scores = data.get('cvss_v2_base_score', set())
                if not any(float(score) >= cvss_v2_base_filter for score in cvss_v2_base_scores if score not in [None, 'nan']):
                    continue  # Skip if no CVSS V2 Base Score meets the filter criteria

            # Check CVSS V2 Exploitability Score filter
            if cvss_v2_exploitability_filter is not None:
                cvss_v2_exploitability_scores = data.get('cvss_v2_exploitability_score', set())
                if not any(float(score) >= cvss_v2_exploitability_filter for score in cvss_v2_exploitability_scores if score not in [None, 'nan']):
                    continue  # Skip if no CVSS V2 Exploitability Score meets the filter criteria

            if cvss_v2_impact_filter is not None:
                cvss_v2_impact_scores = data.get('cvss_v2_impact_score', set())
                if not any(float(score) >= cvss_v2_impact_filter for score in cvss_v2_impact_scores if score not in [None, 'nan']):
                    continue  # Skip if no CVSS V2 Impact Score meets the filter criteria

            # Check CVSS V3.1 Base Score filter
            if cvss_v31_base_filter is not None:
                cvss_v31_base_scores = data.get('cvss_v31_base_score', set())
                if not any(float(score) >= cvss_v31_base_filter for score in cvss_v31_base_scores if score not in [None, 'nan']):
                    continue  # Skip if no CVSS V3.1 Base Score meets the filter criteria

            # Check CVSS V3.1 Exploitability Score filter
            if cvss_v31_exploitability_filter is not None:
                cvss_v31_exploitability_scores = data.get('cvss_v31_exploitability_score', set())
                if not any(float(score) >= cvss_v31_exploitability_filter for score in cvss_v31_exploitability_scores if score not in [None, 'nan']):
                    continue  # Skip if no CVSS V3.1 Exploitability Score meets the filter criteria

            # Check CVSS V3.1 Impact Score filter
            if cvss_v31_impact_filter is not None:
                cvss_v31_impact_scores = data.get('cvss_v31_impact_score', set())
                if not any(float(score) >= cvss_v31_impact_filter for score in cvss_v31_impact_scores if score not in [None, 'nan']):
                    continue  # Skip if no CVSS V3.1 Impact Score meets the filter criteria
                
            
            # If all filters are satisfied, add to filtered data
            filtered_cve_data[cve_id] = data

        cve_data = filtered_cve_data

    excel_output_file = os.path.join(PKCT_REPORT_PATH, f'pkct_report_{timestamp}.xlsx')

    save_to_excel(cve_data, results, excel_output_file, args.sections, result_file)
    logger.info(f"Report generated: {excel_output_file}")

    manifest_file = None
    if args.manifest_filename:
        manifest_file = os.path.basename(args.manifest_filename)

    # Generate the HTML report
    html_output_file_path = os.path.join(PKCT_REPORT_PATH, f'pkct_report_{timestamp}.html')
    patch_files_dir = os.path.join(os.getcwd(), 'patch_files')  # Adjust if your patch files are stored elsewhere
    report_date = datetime.now().strftime("%d-%m-%Y")
    generate_html_report(excel_output_file, html_output_file_path, patch_files_dir, result_file, package_name=pack, version=args.version, sections=args.sections, manifest_file = manifest_file, report_date = report_date)
    logger.info(f"HTML report generated: {html_output_file_path}")

    hmi_report_path = os.path.join(settings.PROJECT_DIR, 'hmi', 'CVEHMI','hmiapp', 'media', 'reports', username, project_id, scan_id)
    if not os.path.exists(hmi_report_path):
        os.makedirs(hmi_report_path)

    if args.report_name:
        report_name = args.report_name
    else:
        report_name = f'pkct_report_{scan_id}'

    excel_report = os.path.join(hmi_report_path, f'{report_name}.xlsx') 
    html_report = os.path.join(hmi_report_path, f'{report_name}.html')

    save_to_excel(cve_data, results, excel_report, args.sections, result_file)
    logger.info(f"Excel report generated: {excel_report}")
    generate_html_report(excel_report, html_report, patch_files_dir, result_file, package_name=pack, version=args.version, sections=args.sections, manifest_file = manifest_file, report_date = report_date)
    logger.info(f"HTML report generated: {html_report}")
    generate_download_html(scan_id, username, project_id, excel_report, html_report, hmi_report_path, HMI_REPORT, version=args.version, manifest_file = manifest_file, report_date=report_date, report_name= report_name)

            
if __name__ == "__main__":
    main()
