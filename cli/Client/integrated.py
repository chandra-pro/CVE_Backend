"""
====================
integrated.py
Inetegrated code for PKCT and CVE search and Report Generation
Author: Shubham, Chandramani Kumar
===================
 
"""


import os
import re
import subprocess
import difflib
import requests
import csv
import base64
from git import Repo, exc
from git.exc import GitCommandError
from datetime import datetime
from bs4 import BeautifulSoup
import shutil
import settings
import django
from uuid import UUID
import sys
import logging
import json
import openpyxl
from urllib.parse import quote
from django.utils.timezone import make_naive
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment
import pandas as pd
from openpyxl import load_workbook
import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from hmi.CVEHMI.integrated_report_functions import *

# Initialize Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'settings') 
django.setup()

from serverapp.models import CVE, CVEDescription,MatchString, CPEMatchInNode, CVEReference, MatchString, CPEMatch, CPEMatchInNode, CVSSMetricV2, CVSSMetricV31, CVEWeakness

LOGS_DIR = os.path.join(settings.PROJECT_DIR,'hmi','CVEHMI','hmiapp','media','logs')
HMI_REPORT = os.path.join(settings.PROJECT_DIR,'hmi','CVEHMI','hmiapp','media','reports')
PKCT_REPORT_PATH = settings.PKCT_REPORT_PATH
os.makedirs(LOGS_DIR, exist_ok=True)



# Configure logging
def setup_logging(scan_id):
    log_file_path = os.path.join(LOGS_DIR, f"{scan_id}_scan.log")
    logging.basicConfig(
        filename=log_file_path,
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )

logger = logging.getLogger()

def read_data_from_file(file_path):
    package_data = {}
    current_package = None
    current_version = None
    current_vendor = None
   
    with open(file_path, 'r') as file:
        lines = file.readlines()
        for line in lines:
            line = line.strip()
            if line.startswith("Package Name :"):
                current_package = line.split(":")[1].strip()
                logging.info(f"Processing package: {current_package}")
                if current_package not in package_data:
                    package_data[current_package] = {}
            elif line.startswith("Version :"):
                current_version = line.split(":")[1].strip()
                logging.info(f"Processing version: {current_version}")
                if current_version not in package_data[current_package]:
                    package_data[current_package][current_version] = {'cve_entries': []}
            elif line.startswith("Vendor:"):
                current_vendor = line.split(":")[1].strip()
            elif line.startswith("CVE ID:"):
                parts = line.split(", ")
                cve_id = parts[0].split(":")[1].strip()
                vulnerable = parts[1].split(":")[1].strip()
                package_data[current_package][current_version]['cve_entries'].append((cve_id, vulnerable, current_vendor))
   
    return package_data

###############################################################################
# Fetch CVE ID Details

def fetch_cve_details(cve_ids, include_sections):
    cve_details = {}
   
    for cve_id, vendor_name in cve_ids:
        try:
            cve = CVE.objects.get(id=cve_id)
            published_date = make_naive(cve.published).date()
            cve_details[cve_id] = {
                'CVE ID': cve.id,
                'Source Identifier': cve.source_identifier,
                'Published': make_naive(cve.published),
                'Published Date': published_date,
                'Last Modified': make_naive(cve.last_modified),
                'Vulnerability Status': cve.vuln_status,
                'Vendor Name': vendor_name
            }

            if 'Description' in include_sections:
                descriptions = CVEDescription.objects.filter(cve=cve).values('lang', 'value')
                
                cve_details[cve_id]['Description'] = []
                for desc in descriptions:
                    if desc['lang'] != 'es':  # Exclude descriptions in Spanish
                        
                        cve_details[cve_id]['Description'].append(desc['value'])
            if 'CVSSV2' in include_sections:
                metrics_v2 = CVSSMetricV2.objects.filter(cve=cve).values(
                    'source', 'version', 'vector_string', 'access_vector',
                    'base_score', 'exploitability_score', 'impact_score'
                )
                cve_details[cve_id]['CVSS V2'] = list(metrics_v2)

            if 'CVSSV3.1' in include_sections:
                metrics_v31 = CVSSMetricV31.objects.filter(cve=cve).values(
                    'source', 'version', 'vector_string', 'attack_vector',
                    'privileges_required', 'base_score', 'exploitability_score', 'impact_score'
                )
                cve_details[cve_id]['CVSS V3.1'] = list(metrics_v31)

            if 'Weaknesses' in include_sections:
                weaknesses = CVEWeakness.objects.filter(cve=cve).values('source', 'type', 'description')
                cve_details[cve_id]['Weaknesses'] = [
                    f"Source: {w['source']}, Type: {w['type']}, Description: {w['description']}"
                    for w in weaknesses
                ]
            if 'References' in include_sections:
                references = CVEReference.objects.filter(cve=cve).values('url', 'source', 'tags')
                cve_details[cve_id]['References'] = [
                    f"URL: {r['url']}, Source: {r['source']}, Tags: {r['tags']}"
                    for r in references
                ]

        except CVE.DoesNotExist:
            logging.error(f"CVE ID {cve_id} not found in the database.")

    # Convert aggregated data into a list of dictionaries
    aggregated_details = []
    for cve_id, details in cve_details.items():
        entry = {
            'CVE ID': details['CVE ID'],
            'Package Name': '',
            'Version': '',
            'Vendor Name': details['Vendor Name'],
            'Vulnerability Status': details['Vulnerability Status'],
            'Source Identifier': details['Source Identifier'],
            'Published': details['Published'],
            'Published Date': details['Published Date'],
            'Last Modified': details['Last Modified'],
        }

        if 'Description' in include_sections:
            entry.update({
                
                'Description': ' | '.join(details.get('Description', [])),
            })
        if 'CVSSV2' in include_sections:
            entry.update({
                'CVSS V2 Source': ', '.join([metric['source'] for metric in details.get('CVSS V2', [])]),
                'CVSS V2 Version': ', '.join([metric['version'] for metric in details.get('CVSS V2', [])]),
                'CVSS V2 Vector String': ', '.join([metric['vector_string'] for metric in details.get('CVSS V2', [])]),
                'CVSS V2 Access Vector': ', '.join([metric['access_vector'] for metric in details.get('CVSS V2', [])]),
                'CVSS V2 Base Score': ', '.join([str(metric['base_score']) for metric in details.get('CVSS V2', [])]),
                'CVSS V2 Exploitability Score': ', '.join([str(metric['exploitability_score']) for metric in details.get('CVSS V2', [])]),
                'CVSS V2 Impact Score': ', '.join([str(metric['impact_score']) for metric in details.get('CVSS V2', [])])
            })

        if 'CVSSV3.1' in include_sections:
            entry.update({
                'CVSS V3.1 Source': ', '.join([metric['source'] for metric in details.get('CVSS V3.1', [])]),
                'CVSS V3.1 Version': ', '.join([metric['version'] for metric in details.get('CVSS V3.1', [])]),
                'CVSS V3.1 Vector String': ', '.join([metric['vector_string'] for metric in details.get('CVSS V3.1', [])]),
                'CVSS V3.1 Attack Vector': ', '.join([metric['attack_vector'] for metric in details.get('CVSS V3.1', [])]),
                'CVSS V3.1 Privileges Required': ', '.join([metric['privileges_required'] for metric in details.get('CVSS V3.1', [])]),
                'CVSS V3.1 Base Score': ', '.join([str(metric['base_score']) for metric in details.get('CVSS V3.1', [])]),
                'CVSS V3.1 Exploitability Score': ', '.join([str(metric['exploitability_score']) for metric in details.get('CVSS V3.1', [])]),
                'CVSS V3.1 Impact Score': ', '.join([str(metric['impact_score']) for metric in details.get('CVSS V3.1', [])])
            })
        if 'Weaknesses' in include_sections:
            entry['Weakness Details'] = ' | '.join(details.get('Weaknesses', []))

        if 'References' in include_sections:
            entry['Reference Details'] = ' | '.join(details.get('References', []))

        aggregated_details.append(entry)
   
    return aggregated_details

###############################################################################################
# Prepare Excel File

def prepare_data_for_excel(package_data, include_sections, filters, username, scan_id):
    sheets = {}
    setup_logging(scan_id)
    for package, versions in package_data.items():
        for version, data in versions.items():
            cve_entries = data['cve_entries']
            cve_ids_with_vendor = [(entry[0], entry[2]) for entry in cve_entries]
           
            cve_details = fetch_cve_details(cve_ids_with_vendor, include_sections)

            # Create a DataFrame from the CVE details
            df = pd.DataFrame(cve_details)

            required_columns = ['CVE ID', 'Package Name', 'Version', 'Vendor Name', 'Patch Status', 'Status Detail', 'Patch File URL', 'Vulnerability Status', 'Published Date', 'Last Modified']
            for col in required_columns:
                if col not in df.columns:
                    df[col] = None  # Add the column with None values if it doesn't exist

            # Ensure CVSS columns are present
            cvss_columns = ['CVSS V2 Base Score', 'CVSS V2 Exploitability Score', 'CVSS V2 Impact Score',
                            'CVSS V3.1 Base Score', 'CVSS V3.1 Exploitability Score', 'CVSS V3.1 Impact Score']
            for col in cvss_columns:
                if col not in df.columns:
                    df[col] = None  # Add the column with None values if it doesn't exist

            if df.empty:
                logger.error(f"No data available for package {package}, version {version}")
                continue

            # Handle 'Published' and 'Published Date' columns
            if 'Published' in df.columns:
                df['Published'] = pd.to_datetime(df['Published'])
                df['Published Date'] = df['Published'].dt.date
                #df['Published Date'] = df['Published'].astype(str).str.split(' ').str[0]
            elif 'Published Date' in df.columns:
                df['Published Date'] = pd.to_datetime(df['Published Date']).dt.date
                #df['Published Date'] = df['Published Date'].astype(str).str.split(' ').str[0]
            else:
                logger.error(f"Neither 'Published' nor 'Published Date' column found for package {package}, version {version}")

            # Apply filters
            if 'PublishedDate' in filters and 'Published Date' in df.columns:
                df['Published Date'] = pd.to_datetime(df['Published Date']).dt.date
                df = df[df['Published Date'] >= filters['PublishedDate']]
                
            df['Published Date'] = df['Published Date'].astype(str).str.split(' ').str[0]
            df['Last Modified'] = df['Last Modified'].astype(str).str.split(' ').str[0]

            # Function to check if any score in a comma-separated string meets the criteria
            def check_scores(scores, threshold):
                if pd.isna(scores):
                    return False
                return any(float(score.strip()) >= threshold for score in str(scores).split(',') if score.strip())

            # Apply CVSS filters
            cvss_columns_dict = {
                'CVSSV2Base': 'CVSS V2 Base Score',
                'CVSSV2Exploitability': 'CVSS V2 Exploitability Score',
                'CVSSV2Impact': 'CVSS V2 Impact Score',
                'CVSSV3.1Base': 'CVSS V3.1 Base Score',
                'CVSSV3.1Exploitability': 'CVSS V3.1 Exploitability Score',
                'CVSSV3.1Impact': 'CVSS V3.1 Impact Score'
            }

            for filter_key, column_name in cvss_columns_dict.items():
                if filter_key in filters:
                    if column_name in df.columns:
                        df = df[df[column_name].apply(lambda x: check_scores(x, filters[filter_key]))]
                    else:
                        logger.warning(f"Column '{column_name}' not found in the data. Skipping this filter.")

            if not df.empty and 'CVE ID' in df.columns:
                # Add 'Package Name' and 'Version' as new columns
                df['Package Name'] = package
                df['Version'] = version

                # Select only the specified sections
                #columns_to_include = ['CVE ID', 'Package Name', 'Version', 'Vendor Name', 'Vulnerability Status', 'Published Date', 'Last Modified']
                columns_to_include = ['CVE ID', 'Package Name', 'Version', 'Vendor Name', 'Patch Status', 'Status Detail', 'Patch File URL', 'Vulnerability Status', 'Published Date', 'Last Modified']
                if 'Description' in include_sections:
                    columns_to_include.extend(['Description'])
                if 'CVSSV2' in include_sections:
                    columns_to_include.extend(['CVSS V2 Source', 'CVSS V2 Version', 'CVSS V2 Vector String', 'CVSS V2 Access Vector', 'CVSS V2 Base Score', 'CVSS V2 Exploitability Score', 'CVSS V2 Impact Score'])
                if 'CVSSV3.1' in include_sections:
                    columns_to_include.extend(['CVSS V3.1 Source', 'CVSS V3.1 Version', 'CVSS V3.1 Vector String', 'CVSS V3.1 Attack Vector', 'CVSS V3.1 Privileges Required', 'CVSS V3.1 Base Score', 'CVSS V3.1 Exploitability Score', 'CVSS V3.1 Impact Score'])
                if 'Weaknesses' in include_sections:
                    columns_to_include.append('Weakness Details')
                if 'References' in include_sections:
                    columns_to_include.append('Reference Details')

                # Only include columns that exist in the DataFrame
                columns_to_include = [col for col in columns_to_include if col in df.columns]

                df = df[columns_to_include]

                # Sort DataFrame by 'CVE ID'
                df = df.sort_values(by='CVE ID')

                # Create a unique sheet name (20 chars of package + 8 chars of version)
                short_package = package[:20]
                short_version = version[:8]
                sheet_name = f"{short_package}_{short_version}"
               
                # Ensure unique sheet names by appending a counter if necessary
                counter = 1
                original_sheet_name = sheet_name
                while sheet_name in sheets:
                    sheet_name = f"{original_sheet_name}_{counter}"
                    counter += 1
               
                sheets[sheet_name] = df
            else:
                logger.warning(f"No data available for package {package}, version {version} after applying filters.")
   
    return sheets

#################################################################################################
# Write CVE Details to Excel file

def write_to_excel(sheets, file_path):
    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
        for sheet_name, df in sheets.items():
            # Ensure the columns are in the correct order
            columns_order = ['CVE ID', 'Package Name', 'Version', 'Vendor Name', 'Patch Status', 'Status Detail', 'Patch File URL', 'Vulnerability Status', 'Published Date', 'Last Modified']
            # Add any additional columns that might be present
            columns_order.extend([col for col in df.columns if col not in columns_order])
            
            # Reorder the DataFrame columns
            df = df.reindex(columns=columns_order)
            
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
           
            # Adjust column width based on the length of the content
            for idx, col in enumerate(df.columns):
                max_length = max(df[col].astype(str).map(len).max(), len(col))
                adjusted_width = max_length + 2  # Add some extra space
                worksheet.set_column(idx, idx, adjusted_width)

##################################################################################################
# Remove Empty Sheets

def remove_empty_sheets(excel_file):
    wb = load_workbook(excel_file)
    sheets_to_remove = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet.max_row <= 1:  # Only header row or completely empty
            sheets_to_remove.append(sheet_name)
        else:
            has_cve = False
            for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0]:  # Check if CVE ID is not empty
                    has_cve = True
                    break
            if not has_cve:
                sheets_to_remove.append(sheet_name)

    for sheet_name in sheets_to_remove:
        del wb[sheet_name]

    if len(wb.sheetnames) == 0:
        # If all sheets are removed, add an empty sheet with a message
        no_cve_sheet = wb.create_sheet("No_CVEs")
        no_cve_sheet['A1'] = "No CVEs found"  # Adding the message to the first cell

    wb.save(excel_file)
    


##################################################################################################

def read_excel(file_path):
    # Read the Excel file
    excel_data = pd.ExcelFile(file_path)
   
    # Create a dictionary to store data
    data = {}
   
    for sheet_name in excel_data.sheet_names:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
       
        # Extract package and version from sheet name
        package, version = sheet_name.rsplit('_', 1)
       
        # Create a unique key for each package-version combination
        key = f"{package}_{version}"
       
        # Ensure CVSS columns are present
        cvss_columns = ['CVSS V2 Base Score', 'CVSS V2 Exploitability Score', 'CVSS V2 Impact Score',
                        'CVSS V3.1 Base Score', 'CVSS V3.1 Exploitability Score', 'CVSS V3.1 Impact Score']
        for col in cvss_columns:
            if col not in df.columns:
                df[col] = None  # Add the column with None values if it doesn't exist

        # Convert relevant columns to numeric, handling multiple scores
        for col in cvss_columns:
            df[col] = df[col].apply(lambda x: ', '.join([str(score) for score in pd.to_numeric(x.split(','), errors='coerce') if pd.notna(score)]) if isinstance(x, str) else x)

        # Add data to dictionary
        data[key] = df
   
    return data

##############################################################################################
# Update Excel File With Patch Status Of CVE Ids

def update_excel_with_patch_info(excel_file_path, txt_file_path):
    """
    Update the Excel file in-place with Patch Status, Status Detail, and Patch File URL columns
    based on the CVE information from the provided text file. Only populate these columns
    for CVE IDs present in both the Excel file and the text file.
    Args:
    - excel_file_path (str): Path to the input Excel file with CVE data (will be updated in-place).
    - txt_file_path (str): Path to the text file containing patch status information.
    """
    
    # Read the TXT file and extract information into a dictionary
    cve_info = {}
    with open(txt_file_path, 'r') as txt_file:
        for line in txt_file:
            match = re.match(r'^(CVE-\d{4}-\d+)\s+(.+?)(?:\s+\(([^)]+)\))?(?:\s+(http[^\s]+))?$', line.strip())
            if match:
                cve_id, patch_status, status_detail, patch_url = match.groups()
                cve_info[cve_id] = {
                    'Patch Status': patch_status,
                    'Status Detail': status_detail if status_detail else 'Not Available',
                    'Patch File URL': patch_url if patch_url else 'Not Available'
                }

    # Load the workbook
    workbook = openpyxl.load_workbook(excel_file_path)

    # Iterate over all sheets in the Excel file
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Read the existing sheet into a DataFrame
        data = sheet.values
        columns = next(data)[0:]  # Get the first row as columns
        df = pd.DataFrame(data, columns=columns)
        
        # Ensure the columns are in the correct order
        columns_order = ['CVE ID', 'Package Name', 'Version', 'Vendor Name', 'Patch Status', 'Status Detail', 'Patch File URL', 'Vulnerability Status', 'Published Date', 'Last Modified']
        # Add any additional columns that might be present
        columns_order.extend([col for col in df.columns if col not in columns_order])
        
        # Reorder the DataFrame columns
        df = df.reindex(columns=columns_order)
        
        # Update only the rows where CVE IDs are present in both Excel and text file
        for index, row in df.iterrows():
            cve_id = row['CVE ID']
            if cve_id in cve_info:
                df.at[index, 'Patch Status'] = cve_info[cve_id]['Patch Status']
                df.at[index, 'Status Detail'] = cve_info[cve_id]['Status Detail']
                df.at[index, 'Patch File URL'] = cve_info[cve_id]['Patch File URL']
        
        # Clear the sheet and write the updated dataframe, including column names
        sheet.delete_rows(1, sheet.max_row)
        # Write column names
        for c_idx, column_name in enumerate(df.columns, 1):
            sheet.cell(row=1, column=c_idx, value=column_name)
        # Write data
        for r_idx, row in enumerate(df.itertuples(index=False), 2):  # Start from row 2
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
        
        # Adjust column width based on the length of the data
        for idx, column in enumerate(df.columns, 1):
            max_length = max(df[column].astype(str).map(len).max(), len(column)) + 2
            sheet.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = max_length
            
            # Specifically adjust the "Published Date" column width if it exists
            if column == "Published Date":
                # Adjust width specifically for date formatting
                sheet.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = 20  # Example width for dates

    # Save the workbook
    workbook.save(excel_file_path)
    print(f"Excel file updated with patch status information: {excel_file_path}")

####################################################################################
# Filter CVE Id on Patch Status

def filter_excel_on_patch_status(file_path, filter_terms):
    """
    Filters an Excel file based on specified Patch Status terms.
    
    :param file_path: Path to the Excel file
    :param filter_terms: List of Patch Status terms to filter by
    """
    # Load the Excel file
    workbook = load_workbook(filename=file_path)
    

    # Iterate through all sheets
    for sheet_name in workbook.sheetnames:
        # Read the sheet into a pandas DataFrame
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        
        # Check if "Patch Status" column exists
        if "Patch Status" not in df.columns:
            print(f"Warning: 'Patch Status' column not found in sheet '{sheet_name}'. Skipping this sheet.")
            logger.warning(f"'Patch Status' column not found in sheet '{sheet_name}'. Skipping this sheet.")
            continue
        
        # Convert "Patch Status" column to string type
        df["Patch Status"] = df["Patch Status"].astype(str)
        
        # Filter the DataFrame based on the specified terms in "Patch Status" column
        filtered_df = df[df["Patch Status"].str.contains('|'.join(filter_terms), case=False, na=False)]
        
        # Clear the existing sheet, but keep the header row
        sheet = workbook[sheet_name]
        sheet.delete_rows(2, sheet.max_row)  # Start deleting from the second row
        # Write the filtered DataFrame back to the sheet, including headers
        for r_idx, row in enumerate(filtered_df.itertuples(index=False), 2):  # Start from the second row
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
    # Save the changes to the same Excel file
    workbook.save(file_path)
    print(f"Excel file filtered based on Patch Status: {file_path}")
    logger.info(f"Excel file filtered based on Patch Status: {file_path}")

##########################################################################################
# Load Blacklist File

def load_blacklist(file_path,scan_id):
    """
    Load CVE IDs from a blacklist file.
    
    :param file_path: Path to the blacklist file (Excel or CSV).
    :return: Set of CVE IDs to be excluded.
    """
    setup_logging(scan_id)
    blacklist = set()
    if file_path.endswith('.xlsx'):
        df = pd.read_excel(file_path, header=None)  # Specify no header
        blacklist = set(df.iloc[:, 0].dropna().astype(str).str.strip().str.upper().tolist())
    elif file_path.endswith('.csv'):
        df = pd.read_csv(file_path, header=None)  # Specify no header
        blacklist = set(df.iloc[:, 0].dropna().astype(str).str.strip().str.upper().tolist())
    else:
        logger.error("Unsupported file format. Please provide an Excel or CSV file.")
    
    return blacklist 

#############################################################################################

def check_scores(score, threshold):
    if pd.isna(score):
        return False
    scores = str(score).split(',')
    return any(float(s) >= threshold for s in scores if s.strip())


##############################################################################################
# Run cve_search_manifest.py and pkct_main.py in parallel execution

def run_script(script_name, *args):
    """Runs the specified script with given arguments and returns the output path."""
    # Run the script with provided arguments
    if None in args:
        raise ValueError("One or more required arguments are None.")
    
    command = ['python3', script_name] + list(args)
    logger.info(f"Executing command: {' '.join(command)}")
    print(f"Executing command: {' '.join(command)}")  # Debugging output
    result = subprocess.run(command, capture_output=True, text=True)
    
    # If successful, extract and return the path
    if result.returncode == 0:
        # Use regex to extract the file path in the result output
        match = re.search(r"Results have been written to (.+)", result.stdout)
        if match:
            output_path = match.group(1).strip()
            return output_path
    else:
        logger.error(f"Error running {script_name}: {result.stderr.strip()}")
        print(f"Error running {script_name}: {result.stderr.strip()}")

    return None
    
############################################################################################

def main():
    # Set up argument parser for Integrated.py
    parser = argparse.ArgumentParser(description="Run pkct_kcc.py and cve_search_manifest.py in parallel.")
    
    # Arguments for pkct_kcc.py
    parser.add_argument('-gk', '--githubkernel', help='GitHub link to user kernel repository', required=False)
    parser.add_argument('-gb', '--githubbranch', help='Branch for the GitHub user kernel repository', required=False)
    parser.add_argument('-db', '--dotbranch', help='Upstream dot kernel branch', required=False)
    parser.add_argument('-u', '--user', help='Original user name', required=False)
    parser.add_argument('-ub', '--upstreambranch', help='upstream kernel branch', required=False)
    parser.add_argument('-pname', '--projectname', help='Project name', required=False)
    parser.add_argument('-build', '--buildpath', help='Build file ', required=True)
    parser.add_argument("-m", "--manifest_filename", help="The filename of the manifest located in the 'sample_manifest' folder.", required=False)
    parser.add_argument("-s", "--sections", nargs='*', help="Additional sections to include in the report (e.g., Description, CVSSV2, CVSSV3.1, Weaknesses, References)")
    parser.add_argument('--filter', nargs='+', help='Filters to apply (PublishedDate 20-05-2022, CVSSV2Base 2.0, CVSSV2Exploitability 1.5, CVSSV2Impact 1.5, CVSSV3.1Base 1.6, CVSSV3.1Exploitability 2.2, CVSSV3.1Impact 2.2)')
    parser.add_argument('-bl', '--blacklist', help='Path to the CSV or XLSX file containing CVE IDs to exclude', required=False)
    parser.add_argument('--username', help='request.user.username', required=False)
    parser.add_argument('--project_id', help='project_id', required=False)
    parser.add_argument('--scan_id', help='scan id', required=False)
    parser.add_argument('--report_name', help='The name of the report to generate')

    args = parser.parse_args()
    username = args.username
    scan_id = args.scan_id
    project_id = args.project_id
    setup_logging(scan_id)

    # Prepare arguments for each script
    pkct_args = [
      '-gk',args.githubkernel,
      '-gb',args.githubbranch,
      '-db',args.dotbranch,      
      '-build',args.buildpath,
      '-u',args.user,
      '-pname',args.projectname,
      '-m',args.manifest_filename,  
      '--username', args.username,
      '--scan_id', args.scan_id
    ]

    cve_args = ['-m', args.manifest_filename] if args.manifest_filename else []
    cve_args.extend(['--username', args.username])
    cve_args.extend(['--scan_id', args.scan_id])

    if args.blacklist:
        cve_args.extend(['--blacklist', args.blacklist])


    # Dictionary to hold output paths
    output_paths = {}

    with ThreadPoolExecutor() as executor:
        futures = {
            executor.submit(run_script, 'pkct_main.py', *pkct_args): 'pkct_main.py',
            executor.submit(run_script, 'cve_search_manifest.py', *cve_args): 'cve_search_manifest.py'
        }
        for future in as_completed(futures):
            script = futures[future]
            output_path = future.result()
            output_paths[script] = output_path

    # Print the results
    for script, path in output_paths.items():
        if path:
            print(f"{script} output path: {path}")
        else:
            print(f"{script} did not generate a valid output path.")

    # Explicitly assign output file paths for cve and pkct
    output_file_path_cve = output_paths.get('cve_search_manifest.py')
    output_file_path_pkct = output_paths.get('pkct_main.py')

    include_sections = args.sections
    filters = {}
    patch_status_filter = []

    if args.filter:
        i = 0
        while i < len(args.filter):
            field = args.filter[i]
            i += 1
            if field == 'PatchStatus':
                while i < len(args.filter) and not args.filter[i] in ['PublishedDate', 'CVSSV2Base', 'CVSSV2Exploitability', 'CVSSV2Impact', 'CVSSV3.1Base', 'CVSSV3.1Exploitability', 'CVSSV3.1Impact']:
                    patch_status_filter.append(args.filter[i])
                    i += 1
            elif field == 'PublishedDate':
                filters[field] = datetime.strptime(args.filter[i], '%d-%m-%Y').date()
                i += 1
            elif field in ['CVSSV2Base', 'CVSSV2Exploitability', 'CVSSV2Impact', 'CVSSV3.1Base', 'CVSSV3.1Exploitability', 'CVSSV3.1Impact']:
                filters[field] = float(args.filter[i])
                i += 1
            else:
                logger.error(f"Unsupported filter field: {field}. Supported fields are 'PatchStatus', 'PublishedDate', 'CVSSV2Base', 'CVSSV2Exploitability', 'CVSSV2Impact', 'CVSSV3.1Base', 'CVSSV3.1Exploitability', 'CVSSV3.1Impact'.")
                return
                
    
    
    
    # Generate a timestamped filename for the results
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_date = datetime.now().strftime("%d-%m-%Y")

    # Load the blacklist if provided
    blacklist = set()
    if args.blacklist:
        blacklist = load_blacklist(args.blacklist,scan_id)
        logger.info(f"Loaded blacklist with {len(blacklist)} CVE IDs: {blacklist}") 

    input_txt_path = output_file_path_cve

    output_excel_path_integrated = os.path.join(PKCT_REPORT_PATH, f'integrated_report_{timestamp}.xlsx')
    output_html_path = os.path.join(PKCT_REPORT_PATH, f'integrated_report_{timestamp}.html')

    hmi_report = settings.HMI_REPORT
    hmi_report_path = os.path.join(settings.PROJECT_DIR, 'hmi', 'CVEHMI','hmiapp', 'media','reports',username,project_id,scan_id)
    if not os.path.exists(hmi_report_path):
        os.makedirs(hmi_report_path)

    if args.report_name:
        report_name = args.report_name
    else:
        report_name = f'integrated_report_{scan_id}'

    excel_report = os.path.join(hmi_report_path, f'{report_name}.xlsx') 
    html_report =  os.path.join(hmi_report_path, f'{report_name}.html')

    manifest_file = os.path.basename(args.manifest_filename)
    
    # Define all possible sections
    all_sections = ['Description', 'CVSSV2', 'CVSSV3.1', 'Weaknesses', 'References']
    # Check if sections are provided; if not, include all sections
    include_sections = args.sections if args.sections is not None else all_sections
    package_data = read_data_from_file(input_txt_path)
    sheets = prepare_data_for_excel(package_data, include_sections,filters, username, scan_id)

    if sheets:
        write_to_excel(sheets, output_excel_path_integrated)
        update_excel_with_patch_info(output_excel_path_integrated,output_file_path_pkct)
        logger.info(f'Excel report generated: {output_excel_path_integrated}')
        write_to_excel(sheets, excel_report)
        update_excel_with_patch_info(excel_report, output_file_path_pkct)
        logger.info(f'Excel report generated: {excel_report}')
        # Apply PatchStatus filter if specified
        if patch_status_filter:
            filter_excel_on_patch_status(output_excel_path_integrated, patch_status_filter)
            logger.info(f'Patch Status filter applied to: {output_excel_path_integrated}')
            filter_excel_on_patch_status(excel_report, patch_status_filter)
            logger.info(f'Patch Status filter applied to:, {excel_report}')
            remove_empty_sheets(output_excel_path_integrated)
            remove_empty_sheets(excel_report)

        data = read_excel(output_excel_path_integrated)

        # Convert 'Published' column to datetime and extract date
        for df in data.values():
            if 'Published' in df.columns:
                df['Published'] = pd.to_datetime(df['Published'])
                df['Published Date'] = df['Published'].dt.date
            elif 'Published Date' in df.columns:
                df['Published Date'] = pd.to_datetime(df['Published Date']).dt.date
            else:
                logger.warning("Warning: Neither 'Published' nor 'Published Date' column found in the data.")
        # Step 2: Generate HTML report with filters
        output_html_path = os.path.join(PKCT_REPORT_PATH, f'integrated_report_{timestamp}.html')
        generate_html_report(data, output_html_path, include_sections, filters, manifest_file= manifest_file, report_date=report_date)
        logger.info(f'HTML report generated: {output_html_path}')
        generate_html_report(data, html_report, include_sections, filters, manifest_file = manifest_file, report_date=report_date)
        print(f"HTML report generated: {hmi_report_path}")
        
        
        logger.info(f"Sections included: {include_sections}")
        for field, value in filters.items():
            logger.info(f"Filtered by {field}: {value}")

    else:
        logger.info("No CVEs found matching the specified criteria. Empty report generated.")
        write_empty_excel(output_excel_path_integrated)
        write_empty_excel(excel_report)
        generate_empty_html_report(output_html_path, include_sections)
        generate_empty_html_report(html_report, include_sections)

    generate_download_html(scan_id, username, project_id, excel_report, html_report, hmi_report_path, hmi_report, manifest_file, report_date, report_name)
    

if __name__ == "__main__":
    main()



